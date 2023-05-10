import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook


wb = openpyxl.load_workbook('商品.xlsx')

# 选择活动工作表
ws = wb.active
for col in ws.columns:
    if '参考价' in col[0].value:
        # 打印该列的列名
        #print(col[0].value)     
        #print(col[0].value.replace("(参考价)", ""))
        dwl = col[0].value.replace("(参考价)", "(参考价)")
        dw = col[0].value.replace("(参考价)", "")
        lm = col[0].column_letter
        #print(dwl)
	
# 确定包含参考价的列号
ref_price_cols = []
for col in ws.iter_cols():
    for cell in col:
        if cell.value and '参考价' in str(cell.value):
            ref_price_cols.append(cell.column)

# 打印出每个包含参考价的列中有数据的单元格所在的行和列号
for col in ref_price_cols:
    #print(f'包含参考价的列名称：{ws.cell(row=1, column=col).value}')
    dw = ws.cell(row=1, column=col).value.replace("(参考价)", "")
    #print(dw)
    for row in range(2, ws.max_row+1):
        if ws.cell(row=row, column=col).value:
            #print(f'第 {row} 行，第 {col} 列：{ws.cell(row=row, column=col).value}')
            dj = ws.cell(row=row, column=col).value
            dwdj = (f'{dw}(参考价: {dj}.00 预估成本: 0')
            print(dwdj,end='\n')
