import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook


wb = openpyxl.load_workbook('商品.xlsx')

# 选择活动工作表
ws = wb.active
for col in ws.columns:
    if '参考价' in col[0].value:
        # 打印该列的列名       
        dwl = col[0].value.replace("(参考价)", "(参考价)")
        dw = col[0].value.replace("(参考价)", "")
        lm = col[0].column_letter
	
# 创建字典，保存每行不同列的参考价数据
ref_price_rows = {}
for col in ws.iter_cols():
    for cell in col:
        if cell.value and '参考价' in str(cell.value):
            col_num = cell.column
            dw = cell.value.replace("(参考价)", "")
            for row in range(2, ws.max_row+1):
                if ws.cell(row=row, column=col_num).value:
                    if row not in ref_price_rows:
                        ref_price_rows[row] = {}
                    ref_price_rows[row][col_num] = ws.cell(row=row, column=col_num).value

# 将同一行的参考价数据合并输出
for row in sorted(ref_price_rows.keys()):
    result = f"第{row}行: "
    for col_num in sorted(ref_price_rows[row].keys()):
        col_name = ws.cell(row=1, column=col_num).value.replace("(参考价)", "")
        ref_price = ref_price_rows[row][col_num]
        result += f"{col_name}(参考价: {ref_price}.00 预估成本: 0) "
    print(result)
    
df = pd.DataFrame(result)

# 写入Excel文件
with pd.ExcelWriter('结果.xlsx') as writer:
    df.to_excel(writer, sheet_name='结果', index=False)