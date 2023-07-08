from ctypes.wintypes import WPARAM
from tkinter import N
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
import os
import zipfile
from bs4 import BeautifulSoup
import glob
import tkinter as tk
from tkinter import filedialog
import shutil
# 创建一个文件夹，如果不存在择创建 反之跳过
importfile = "已生成的导入模板"
if not os.path.exists(importfile):
    os.makedirs(importfile)
else:
    print(f"{importfile}文件夹已经存在，已直接使用。")

# 获取当前工作目录路径
current_dir = os.getcwd()

# 查找当前目录下所有Excel文件
excel_files = glob.glob(os.path.join(current_dir, "*.xlsx"))

# 如果没有找到Excel文件，输出提示信息并结束程序
if not excel_files:
    print("当前目录下没有Excel文件！")
    exit()
    
# 让用户选择一个文件
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="请选择从聚食汇云后台导出的excel文件", filetypes=[("Excel文件", "*.xlsx")])

    
# 如果用户没有选择文件，结束程序
if not file_path:
    print("没有选择文件！")
    exit()


# 读取用户选择的Excel文件
# 读取excel文件，假设文件名为data.xlsx
#df = pd.read_excel("data.xlsx")

# 读取D2单元格的值
wb= openpyxl.load_workbook(file_path)
wbdwdj= openpyxl.load_workbook(file_path)

ws = wb.active
wsdwdj = wbdwdj.active

file_name = ws['F2'].value
# 创建新的Excel文件并定义列名
qy = '启用'
qtpx = '前台排序'
ckj1 = '参考价1'
hyj1 = '会员价1'
zjm = '助记码'
bm ='别名'
pp = '品牌'
qyzf = '启用做法'
spzf = '商品做法'
spzflx = '商品做法类型'
tcsp = '套餐商品'
blzk = '比例折扣'
sjsp = '时价商品'
lssp = '临时商品'
jslrsl = '即时录入数量'
cz = '称重'
yxsydjj = '允许使用代金券'
yxsykye = '允许使用卡余额'
tssp = '特色商品'
xzdcb = '下载点菜宝'
sfwf = '收服务费'
yxjf = '允许积分'
jrzdxf = '计入最低消费'
new_excel = pd.DataFrame(columns=['编号', '名称', '所属小类', '所属大类', '商品类型', '单位', '启用', '前台排序', '参考价1', '会员价1', '助记码', '别名', '品牌', '启用做法', '商品做法', '商品做法类型', '套餐商品', '比例折扣', '时价商品', '临时商品', '即时录入数量', '称重', '允许使用代金券', '允许使用卡余额', '特色商品', '下载点菜宝', '收服务费', '允许积分', '计入最低消费'])

# 从指定Excel文件中读取数据并选择需要的列
old_excel = pd.read_excel(file_path, sheet_name='商品', 
                          usecols=['编号', '名称', '所属小类', '所属大类', '商品类型','前台排序','套餐商品'],
                          dtype={'编号': str})

ckj_excel = pd.read_excel(file_path, sheet_name='商品')
df_ref = ckj_excel.filter(like='参考价')
ref_cols = ckj_excel.filter(like='参考价').columns.tolist()

ref_price_rows = {}
for col in ws.iter_cols():
    for cell in col:
        if cell.value and '参考价' in str(cell.value):
            col_num = cell.column
            dw = cell.value.replace("(参考价)", "")
            for row in range(2, ws.max_row+1):
                if ws.cell(row=row, column=col_num).value:
                    if row not in ref_price_rows:
                        ref_price_rows[row] = {}
                    ref_price_rows[row][col_num] = ws.cell(row=row, column=col_num).value

# 打印出每个包含参考价的列中有数据的单元格所在的行和列号，并将其添加到new_excel['单位']中
for row in sorted(ref_price_rows.keys()):
    result = ''
    for col_num in sorted(ref_price_rows[row].keys()):
        col_name = ws.cell(row=1, column=col_num).value.replace("(参考价)", "")
        ref_price = ref_price_rows[row][col_num]
        result += f"{col_name}(参考价:{ref_price}.00 预估成本:0.00)," # 将每个单元格的数据后面添加一个'、'符号
    new_excel.at[row-2, '单位'] = result.rstrip(',') # 将最右侧的'、'符号去掉，并将数据赋值给new_excel['单位']的对应行


# 使用生成前台排序从1开始 想关闭请注释(new_excel.at[i - 1, '前台排序'] = i_str)取消注释(new_excel[qtpx] = old_excel['前台排序'])
start_number = 1
end_number = ws.max_row - 0  # 减去标题行为有效数据行数

for i, row in enumerate(ws.iter_rows(min_row=2), start=start_number):
    row[0].value = i
    i_str = str(i)
    new_excel.at[i - 1, '前台排序'] = i_str


#份(参考价:12.00 预估成本:0.00),两份(参考价:18.00 预估成本:0.00)
# 将读取到的数据添加到新Excel文件中
new_excel['编号'] = old_excel['编号']
new_excel['名称'] = old_excel['名称']
new_excel['所属小类'] = old_excel['所属小类']
new_excel['所属大类'] = old_excel['所属大类']
new_excel['商品类型'] = old_excel['商品类型']
new_excel[qy] = '1'
#new_excel[qtpx] = old_excel['前台排序'] #获取原来的排序
new_excel[ckj1] = ''
new_excel[hyj1] = ''
new_excel[zjm] = ''
new_excel[bm] = ''
new_excel[pp] = ''
new_excel[qyzf] = ''
new_excel[spzf] = ''
new_excel[tcsp] = ''
new_excel[spzflx] = ''
new_excel[tcsp] = old_excel['套餐商品']
new_excel[blzk] = '1'
new_excel[sjsp] = ''
new_excel[lssp] = ''
new_excel[jslrsl] = ''
new_excel[yxsydjj] = '1'
new_excel[yxsykye] = '1'
new_excel[tssp] = ''
new_excel[xzdcb] = '1'
new_excel[sfwf] = '1'
new_excel[yxjf] = '1'
new_excel[jrzdxf] = '1'

# 将数据写入到新的Excel文件中
writer = pd.ExcelWriter(f'{importfile}/{file_name}.xlsx', engine='openpyxl')
writer.book = Workbook() # 新建excel文件
new_excel.to_excel(writer, sheet_name='商品', index=False) # 写入数据
writer.save()

