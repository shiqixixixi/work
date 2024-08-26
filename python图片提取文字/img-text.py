import pytesseract
import openpyxl
from PIL import Image, ImageFile
ImageFile.LOAD_TRUNCATED_IMAGES = True

img = Image.open('food.jpg').resize((16850, 5906))
# 打开图片文件
#img = Image.open('')
# 使用pytesseract进行OCR识别
text = pytesseract.image_to_string(img, lang='eng')
# 将识别结果按行分割
lines = text.split('\n')

# 创建Excel工作簿
#workbook = openpyxl.Workbook()
workbook = openpyxl.Workbook(encoding='utf-8')
sheet = workbook.active

# 遍历行，将菜名和价格分别写入Excel
for i in range(len(lines)):
    # 如果行中包含数字，则认为是价格
    if any(char.isdigit() for char in lines[i]):
        sheet.cell(row=i+1, column=2, value=lines[i])
        #sheet.cell(row=i+1, column=1, value=lines[i].encode('utf-8')) 
    # 否则认为是菜名
    else:
        sheet.cell(row=i+1, column=1, value=lines[i])
        #sheet.cell(row=i+1, column=2, value=lines[i].encode('utf-8'))

# 保存Excel文件
workbook.save('food.xlsx')
